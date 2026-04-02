"""
Excel Manager — reads and writes jobs.xlsx using openpyxl.
Handles deduplication (by URL), header creation, and row appending.
"""

from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from src.config import EXCEL_PATH, EXCEL_COLUMNS
from src.job_matcher import JobMatch


def _create_workbook(path: Path) -> openpyxl.Workbook:
    """Create a new workbook with styled headers."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Jobs"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, col_name in enumerate(EXCEL_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Set column widths
    col_widths = {
        "Job Title": 35,
        "Company": 25,
        "Location": 20,
        "Source": 15,
        "URL": 50,
        "Date Posted": 14,
        "Match Score": 13,
        "Matched Skills": 40,
        "Missing Skills": 30,
        "Salary": 20,
        "Date Found": 14,
        "Status": 12,
        "Notes": 30,
    }
    for col_idx, col_name in enumerate(EXCEL_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_name, 20)

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"
    wb.save(path)
    print(f"[excel] Created new workbook at {path}")
    return wb


def _load_existing_urls(ws) -> set[str]:
    """Read all existing job URLs from the sheet (column 5 = URL)."""
    urls = set()
    url_col = EXCEL_COLUMNS.index("URL") + 1
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[url_col - 1]:
            urls.add(str(row[url_col - 1]).strip())
    return urls


def _score_fill(score: int) -> PatternFill:
    """Return a colour fill based on match score."""
    if score >= 80:
        color = "C6EFCE"  # green
    elif score >= 65:
        color = "FFEB9C"  # yellow
    else:
        color = "FFCCCC"  # light red
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def append_jobs(matches: list[JobMatch]) -> int:
    """
    Append new job matches to jobs.xlsx.
    Skips duplicates (by URL). Returns the count of newly added rows.
    """
    path = EXCEL_PATH

    if path.exists():
        wb = openpyxl.load_workbook(path)
        ws = wb.active
    else:
        wb = _create_workbook(path)
        ws = wb.active

    existing_urls = _load_existing_urls(ws)
    today = date.today().isoformat()
    added = 0

    score_col_idx = EXCEL_COLUMNS.index("Match Score") + 1
    url_col_idx = EXCEL_COLUMNS.index("URL") + 1

    for match in matches:
        url = match.listing.url.strip()
        if url in existing_urls:
            continue  # deduplicate

        row_data = {
            "Job Title": match.listing.title,
            "Company": match.listing.company,
            "Location": match.listing.location,
            "Source": match.listing.source,
            "URL": url,
            "Date Posted": match.listing.date_posted,
            "Match Score": match.score,
            "Matched Skills": match.matched_skills,
            "Missing Skills": match.missing_skills,
            "Salary": match.listing.salary,
            "Date Found": today,
            "Status": "New",
            "Notes": match.reason,
        }

        row_values = [row_data.get(col, "") for col in EXCEL_COLUMNS]
        next_row = ws.max_row + 1
        ws.append(row_values)

        # Color the score cell
        score_cell = ws.cell(row=next_row, column=score_col_idx)
        score_cell.fill = _score_fill(match.score)

        # Make URL a hyperlink
        url_cell = ws.cell(row=next_row, column=url_col_idx)
        url_cell.hyperlink = url
        url_cell.font = Font(color="0563C1", underline="single")

        # Wrap text for skills columns
        for col_name in ("Matched Skills", "Missing Skills", "Notes"):
            col_idx = EXCEL_COLUMNS.index(col_name) + 1
            ws.cell(row=next_row, column=col_idx).alignment = Alignment(wrap_text=True)

        existing_urls.add(url)
        added += 1

    if added > 0:
        wb.save(path)
        print(f"[excel] Appended {added} new jobs → {path}")
    else:
        print("[excel] No new jobs to append (all duplicates)")

    return added


def get_stats() -> dict:
    """Return basic stats about the jobs.xlsx file."""
    path = EXCEL_PATH
    if not path.exists():
        return {"total": 0, "new": 0, "applied": 0, "path": str(path)}

    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active

    status_col = EXCEL_COLUMNS.index("Status") + 1
    total = 0
    status_counts: dict[str, int] = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        total += 1
        status = str(row[status_col - 1] or "New")
        status_counts[status] = status_counts.get(status, 0) + 1

    return {
        "total": total,
        "new": status_counts.get("New", 0),
        "applied": status_counts.get("Applied", 0),
        "path": str(path),
        "status_counts": status_counts,
    }
